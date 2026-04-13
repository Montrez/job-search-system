#!/usr/bin/env python3
"""
seed_jobs.py — Populate contacts.json + Excel with AI-generated job leads.
Provider priority: Groq (free, fast) → Gemini → built-in fallback leads.

Get a free Groq key at: https://console.groq.com
Add to .env:  GROQ_API_KEY=gsk_...
"""
import json
import os
import sys
import urllib.request
import urllib.error
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent
DATA_FILE     = BASE_DIR / "data" / "contacts.json"
CRITERIA_FILE = BASE_DIR / "data" / "criteria.json"
EXCEL_FILE    = BASE_DIR / "Job_Search_Tracker_Montrez.xlsx"
GEMINI_URL    = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
GROQ_URL      = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL    = "llama-3.3-70b-versatile"

DEFAULT_CRITERIA = {
    "worktype": "Remote Only",
    "minsalary": "$80,000",
    "roles": "Sales Engineer, Solutions Architect, Pre-Sales, Consulting",
    "geo": "US-based remote preferred",
    "size": "Small (1–100)",
    "target": "100",
    "musthaves": "Remote flexibility, base salary above $80k, US resident allowed",
    "nicetohaves": "Healthcare, 401k match, equity, flexible hours",
}


def get_env_key(name: str) -> str:
    val = os.getenv(name, "")
    if val:
        return val
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith(f"{name}="):
                return line.split("=", 1)[1].strip()
    return ""


def call_groq(prompt: str, key: str) -> str:
    payload = {
        "model": GROQ_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.8,
        "max_tokens": 4096,
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        GROQ_URL,
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {key}",
            "User-Agent": "Mozilla/5.0 (compatible; job-search-bot/1.0)",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data["choices"][0]["message"]["content"]


def load_criteria() -> dict:
    if CRITERIA_FILE.exists():
        try:
            return json.loads(CRITERIA_FILE.read_text())
        except Exception:
            pass
    return DEFAULT_CRITERIA


def call_gemini(prompt: str, key: str) -> str:
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.8, "maxOutputTokens": 4096},
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"{GEMINI_URL}?key={key}",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data["candidates"][0]["content"]["parts"][0]["text"]


FALLBACK_LEADS = [
    {"company": "Axiom Analytics", "contact": "Sarah Mitchell", "email": "s.mitchell@axiomanalytics.io", "linkedin": "https://linkedin.com/in/sarah-mitchell-rec", "jobtitle": "Sales Engineer", "joburl": "https://axiomanalytics.io/careers/sales-engineer", "location": "remote", "salary": "$95,000 - $125,000", "status": "not_contacted", "notes": "Series B data analytics SaaS. Strong technical sales team, mentions competitive equity."},
    {"company": "Veridian Cloud", "contact": "Marcus Thompson", "email": "m.thompson@veridiancloud.com", "linkedin": "https://linkedin.com/in/marcus-thompson-tech", "jobtitle": "Solutions Architect", "joburl": "https://veridiancloud.com/jobs/solutions-architect", "location": "remote", "salary": "$110,000 - $140,000", "status": "not_contacted", "notes": "Cloud infra startup, 80 employees. Mentioned pre-sales demos to Fortune 500s."},
    {"company": "Nuvelo Security", "contact": "Priya Chandran", "email": "priya.chandran@nuvelosec.com", "linkedin": "https://linkedin.com/in/priya-chandran-recruiter", "jobtitle": "Pre-Sales Engineer", "joburl": "https://nuvelosec.com/careers/presales", "location": "remote", "salary": "$90,000 - $115,000", "status": "not_contacted", "notes": "Cybersecurity SaaS, 60 employees. Strong benefits package mentioned in JD."},
    {"company": "Lattice Systems", "contact": "David Reyes", "email": "d.reyes@latticesystems.co", "linkedin": "https://linkedin.com/in/david-reyes-hr", "jobtitle": "Solutions Engineer", "joburl": "https://latticesystems.co/jobs/se", "location": "remote", "salary": "$100,000 - $130,000", "status": "not_contacted", "notes": "HR Tech platform, remote-first culture. Engineering-led sales motion."},
    {"company": "Phalanx Data", "contact": "Jennifer Wu", "email": "j.wu@phalanxdata.com", "linkedin": "https://linkedin.com/in/jennifer-wu-talent", "jobtitle": "Technical Sales Consultant", "joburl": "https://phalanxdata.com/careers/tsc", "location": "remote", "salary": "$85,000 - $110,000", "status": "not_contacted", "notes": "Data pipeline tooling, 45 employees. Founders from Databricks/Snowflake."},
    {"company": "Crestview DevOps", "contact": "Tyler Brooks", "email": "t.brooks@crestviewdevops.io", "linkedin": "https://linkedin.com/in/tyler-brooks-recruiter", "jobtitle": "Sales Engineer", "joburl": "https://crestviewdevops.io/jobs/sales-engineer", "location": "remote", "salary": "$92,000 - $120,000", "status": "not_contacted", "notes": "DevOps toolchain SaaS, 55 employees. Very active recruiting, fast interview process."},
    {"company": "Meridian AI", "contact": "Ashley Foster", "email": "a.foster@meridian.ai", "linkedin": "https://linkedin.com/in/ashley-foster-meridian", "jobtitle": "Solutions Architect", "joburl": "https://meridian.ai/careers/sa", "location": "remote", "salary": "$120,000 - $150,000", "status": "not_contacted", "notes": "AI/ML platform startup, 90 employees. Series C, strong growth trajectory."},
    {"company": "Clearpath Fintech", "contact": "Nathan Cole", "email": "n.cole@clearpath.finance", "linkedin": "https://linkedin.com/in/nathan-cole-fintech", "jobtitle": "Pre-Sales Engineer", "joburl": "https://clearpath.finance/jobs/presales-engineer", "location": "remote", "salary": "$88,000 - $112,000", "status": "not_contacted", "notes": "Payments infrastructure, 70 employees. Technical sales to CFOs and finance teams."},
    {"company": "Spectra Ops", "contact": "Monica Garza", "email": "m.garza@spectraops.com", "linkedin": "https://linkedin.com/in/monica-garza-spectra", "jobtitle": "Sales Engineer", "joburl": "https://spectraops.com/careers/se-remote", "location": "remote", "salary": "$96,000 - $122,000", "status": "not_contacted", "notes": "IT Ops automation, 65 employees. Prefers candidates with enterprise demo experience."},
    {"company": "Orion Platform", "contact": "James Nguyen", "email": "j.nguyen@orionplatform.io", "linkedin": "https://linkedin.com/in/james-nguyen-orion", "jobtitle": "Solutions Engineer", "joburl": "https://orionplatform.io/jobs/solutions-engineer", "location": "remote", "salary": "$105,000 - $135,000", "status": "not_contacted", "notes": "API management platform, 85 employees. Great Glassdoor reviews for work-life balance."},
    {"company": "Apex Revenue", "contact": "Chloe Kim", "email": "c.kim@apexrevenue.com", "linkedin": "https://linkedin.com/in/chloe-kim-apex", "jobtitle": "Technical Sales Consultant", "joburl": "https://apexrevenue.com/careers/tsc", "location": "remote", "salary": "$90,000 - $118,000", "status": "not_contacted", "notes": "Sales intelligence SaaS, 50 employees. Good comp structure + uncapped variable."},
    {"company": "Terraform Labs", "contact": "Brandon Scott", "email": "b.scott@terraformlabs.io", "linkedin": "https://linkedin.com/in/brandon-scott-tf", "jobtitle": "Pre-Sales Engineer", "joburl": "https://terraformlabs.io/jobs/presales", "location": "remote", "salary": "$93,000 - $119,000", "status": "not_contacted", "notes": "Infrastructure automation, 75 employees. Fast-growing, Series B closed $40M."},
    {"company": "Vantage Analytics", "contact": "Sophia Reed", "email": "s.reed@vantageanalytics.com", "linkedin": "https://linkedin.com/in/sophia-reed-vantage", "jobtitle": "Sales Engineer", "joburl": "https://vantageanalytics.com/careers/sales-eng", "location": "remote", "salary": "$98,000 - $128,000", "status": "not_contacted", "notes": "BI tooling for SMBs, 60 employees. Strong referral culture."},
    {"company": "Luminary Health Tech", "contact": "Daniel Park", "email": "d.park@luminaryhealth.io", "linkedin": "https://linkedin.com/in/daniel-park-lht", "jobtitle": "Solutions Architect", "joburl": "https://luminaryhealth.io/jobs/sa", "location": "remote", "salary": "$115,000 - $145,000", "status": "not_contacted", "notes": "Healthcare data SaaS, 95 employees. HIPAA-compliant platform, interesting domain."},
    {"company": "Cascade Commerce", "contact": "Rachel Evans", "email": "r.evans@cascadecommerce.com", "linkedin": "https://linkedin.com/in/rachel-evans-cascade", "jobtitle": "Pre-Sales Engineer", "joburl": "https://cascadecommerce.com/careers/presales", "location": "remote", "salary": "$87,000 - $112,000", "status": "not_contacted", "notes": "E-commerce platform for D2C brands, 55 employees. Lots of product demos needed."},
    {"company": "Nexus Cyber", "contact": "Kevin Torres", "email": "k.torres@nexuscyber.com", "linkedin": "https://linkedin.com/in/kevin-torres-nexus", "jobtitle": "Sales Engineer", "joburl": "https://nexuscyber.com/jobs/sales-engineer", "location": "remote", "salary": "$100,000 - $130,000", "status": "not_contacted", "notes": "Zero-trust security SaaS, 80 employees. Strong technical depth required."},
    {"company": "Pulsar Automation", "contact": "Laura Jensen", "email": "l.jensen@pulsarautomation.io", "linkedin": "https://linkedin.com/in/laura-jensen-pulsar", "jobtitle": "Solutions Engineer", "joburl": "https://pulsarautomation.io/jobs/se", "location": "remote", "salary": "$97,000 - $125,000", "status": "not_contacted", "notes": "RPA + workflow automation, 70 employees. SMB-to-mid-market focus."},
    {"company": "Flux Consulting Group", "contact": "Patrick Murphy", "email": "p.murphy@fluxconsulting.com", "linkedin": "https://linkedin.com/in/patrick-murphy-flux", "jobtitle": "Technical Sales Consultant", "joburl": "https://fluxconsulting.com/careers/tsc-remote", "location": "remote", "salary": "$91,000 - $116,000", "status": "not_contacted", "notes": "Digital transformation consulting, 40 employees. Remote-first since founding."},
    {"company": "Cobalt DevTools", "contact": "Samantha Hill", "email": "s.hill@cobaltdevtools.com", "linkedin": "https://linkedin.com/in/samantha-hill-cobalt", "jobtitle": "Sales Engineer", "joburl": "https://cobaltdevtools.com/jobs/se", "location": "remote", "salary": "$94,000 - $120,000", "status": "not_contacted", "notes": "Developer tooling SaaS, 45 employees. PLG motion transitioning to enterprise sales."},
    {"company": "Nimbus Financial", "contact": "Carlos Rivera", "email": "c.rivera@nimbusfinancial.com", "linkedin": "https://linkedin.com/in/carlos-rivera-nimbus", "jobtitle": "Pre-Sales Engineer", "joburl": "https://nimbusfinancial.com/careers/presales", "location": "remote", "salary": "$89,000 - $114,000", "status": "not_contacted", "notes": "Fintech risk platform, 65 employees. Looking for someone with API/integration background."},
    {"company": "Ethos Data Co", "contact": "Megan Walsh", "email": "m.walsh@ethosdataco.com", "linkedin": "https://linkedin.com/in/megan-walsh-ethos", "jobtitle": "Solutions Architect", "joburl": "https://ethosdataco.com/jobs/solutions-architect", "location": "remote", "salary": "$112,000 - $142,000", "status": "not_contacted", "notes": "Data governance SaaS, 88 employees. Architecture role works closely with CTO office."},
    {"company": "Prism Cloud Partners", "contact": "Alex Donovan", "email": "a.donovan@prismcloudpartners.com", "linkedin": "https://linkedin.com/in/alex-donovan-prism", "jobtitle": "Sales Engineer", "joburl": "https://prismcloudpartners.com/careers/se", "location": "remote", "salary": "$99,000 - $127,000", "status": "not_contacted", "notes": "AWS/GCP managed services, 75 employees. Strong OTE with technical bonus component."},
    {"company": "Harborview SaaS", "contact": "Tiffany Chen", "email": "t.chen@harborviewsaas.com", "linkedin": "https://linkedin.com/in/tiffany-chen-hv", "jobtitle": "Technical Sales Consultant", "joburl": "https://harborviewsaas.com/careers/tsc", "location": "remote", "salary": "$86,000 - $110,000", "status": "not_contacted", "notes": "Project management SaaS, 50 employees. Good culture scores, async-first team."},
    {"company": "Quantum Revenue Ops", "contact": "Josh Patel", "email": "j.patel@quantumrevops.com", "linkedin": "https://linkedin.com/in/josh-patel-qro", "jobtitle": "Solutions Engineer", "joburl": "https://quantumrevops.com/jobs/solutions-eng", "location": "remote", "salary": "$103,000 - $132,000", "status": "not_contacted", "notes": "RevOps automation platform, 70 employees. Very strong leadership team from Salesforce."},
    {"company": "Stellar Integrations", "contact": "Diana Lopez", "email": "d.lopez@stellarintegrations.io", "linkedin": "https://linkedin.com/in/diana-lopez-stellar", "jobtitle": "Pre-Sales Engineer", "joburl": "https://stellarintegrations.io/careers/presales", "location": "remote", "salary": "$91,000 - $117,000", "status": "not_contacted", "notes": "iPaaS / integration platform, 60 employees. Looking for pre-sales with connector/API experience."},
]


def _build_prompt(criteria: dict, count: int) -> str:
    today = date.today().isoformat()
    roles = criteria.get('roles', 'Sales Engineer, Solutions Architect')
    role_list = ", ".join(f'"{r.strip()}"' for r in roles.split(","))
    return f"""You are a job search data generator. Generate {count} realistic job leads for a candidate with these search criteria:
- Roles: {roles}
- Work Type: {criteria.get('worktype')}
- Min Salary: {criteria.get('minsalary')}
- Geography: {criteria.get('geo')}
- Company Size: {criteria.get('size')}
- Must-Haves: {criteria.get('musthaves')}

Generate {count} distinct job leads as a JSON array. Each entry must have these exact fields:
{{
  "company": "Real-sounding company name (tech/SaaS/fintech/consulting - small to mid size)",
  "contact": "First Last (recruiter or hiring manager name)",
  "email": "firstname.lastname@company.com (realistic)",
  "linkedin": "https://linkedin.com/in/firstname-lastname-abc123",
  "jobtitle": "One of: {role_list}",
  "joburl": "https://company.com/careers/job-id or https://linkedin.com/jobs/view/123456789",
  "location": "remote",
  "salary": "$XX,000 - $XX,000 (realistic range above $80k, up to $150k)",
  "status": "not_contacted",
  "notes": "1-2 sentence note about why this role is interesting or what to mention in outreach"
}}

Use real-sounding but fictional companies in: SaaS, DevTools, Fintech, Cybersecurity, Data/Analytics, Cloud Infrastructure, HR Tech, Sales Tech.
Vary seniority (mid-level $85k-$110k, senior $110k-$145k). Today: {today}.
Return ONLY the JSON array, no markdown fences, no extra text."""


def _parse_leads(raw: str) -> list:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        raw = raw.rsplit("```", 1)[0]
    leads = json.loads(raw.strip())
    print(f"  Generated {len(leads)} leads.")
    return leads


def generate_job_leads_groq(criteria: dict, key: str, count: int = 25) -> list:
    print(f"  Calling Groq ({GROQ_MODEL})...")
    return _parse_leads(call_groq(_build_prompt(criteria, count), key))


def generate_job_leads(criteria: dict, key: str, count: int = 25) -> list:
    print("  Calling Gemini (gemini-2.0-flash)...")
    return _parse_leads(call_gemini(_build_prompt(criteria, count), key))


def build_contacts(leads: list) -> list:
    today = datetime.now()
    contacts = []
    for i, lead in enumerate(leads):
        days_offset = i % 7
        followup = (today + timedelta(days=3 + days_offset)).date().isoformat()
        record = {
            "id":          str(int(today.timestamp() * 1000) + i),
            "company":     lead.get("company", ""),
            "contact":     lead.get("contact"),
            "email":       lead.get("email"),
            "linkedin":    lead.get("linkedin"),
            "jobtitle":    lead.get("jobtitle"),
            "joburl":      lead.get("joburl"),
            "location":    lead.get("location", "remote"),
            "salary":      lead.get("salary"),
            "status":      lead.get("status", "not_contacted"),
            "followup":    followup,
            "confirmed":   "false",
            "submittedby": None,
            "notes":       lead.get("notes"),
            "dateAdded":   today.isoformat(),
            "lastUpdated": today.isoformat(),
        }
        contacts.append(record)
    return contacts


def save_contacts(contacts: list) -> None:
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    DATA_FILE.write_text(json.dumps(contacts, indent=2, default=str))
    print(f"  Saved {len(contacts)} contacts → {DATA_FILE}")


def update_excel(contacts: list) -> None:
    if not EXCEL_FILE.exists():
        print(f"  Excel file not found: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Tracker"]

    # Clear existing data rows (row 4 onward), keep header rows 1-3
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Color definitions
    header_fill  = PatternFill("solid", fgColor="1a1d27")
    even_fill    = PatternFill("solid", fgColor="22263a")
    odd_fill     = PatternFill("solid", fgColor="1a1d27")
    status_fills = {
        "not_contacted":       PatternFill("solid", fgColor="4A5568"),
        "email_sent":          PatternFill("solid", fgColor="2B6CB0"),
        "responded":           PatternFill("solid", fgColor="B7791F"),
        "interview_scheduled": PatternFill("solid", fgColor="553C9A"),
        "submitted":           PatternFill("solid", fgColor="9C4221"),
        "offer":               PatternFill("solid", fgColor="276749"),
        "rejected":            PatternFill("solid", fgColor="9B2C2C"),
    }

    col_map = {
        "company":     2,
        "contact":     3,
        "email":       4,
        "linkedin":    5,
        "jobtitle":    6,
        "joburl":      7,
        "location":    8,
        "salary":      9,
        "dateAdded":   10,
        "status":      11,
        "followup":    17,
        "confirmed":   18,
        "submittedby": 19,
        "notes":       20,
    }

    for i, contact in enumerate(contacts):
        row_num = i + 4
        fill = even_fill if i % 2 == 0 else odd_fill

        # Row number formula
        ws.cell(row=row_num, column=1).value = i + 1

        for field, col in col_map.items():
            val = contact.get(field)
            cell = ws.cell(row=row_num, column=col)
            if field == "dateAdded" and val:
                try:
                    cell.value = datetime.fromisoformat(val).strftime("%Y-%m-%d")
                except Exception:
                    cell.value = val
            elif field == "status":
                cell.value = val
                cell.fill = status_fills.get(val, fill)
            else:
                cell.value = val
            if field != "status":
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="center")

        # Row height
        ws.row_dimensions[row_num].height = 18

    # Update Dashboard sheet stats
    if "Dashboard" in wb.sheetnames:
        ds = wb["Dashboard"]
        total = len(contacts)
        not_contacted = sum(1 for c in contacts if c["status"] == "not_contacted")
        email_sent    = sum(1 for c in contacts if c["status"] == "email_sent")
        responded     = sum(1 for c in contacts if c["status"] == "responded")
        interviews    = sum(1 for c in contacts if c["status"] == "interview_scheduled")
        offers        = sum(1 for c in contacts if c["status"] == "offer")

        stat_map = {
            (3, 2): total,
            (4, 2): not_contacted,
            (5, 2): email_sent,
            (6, 2): responded,
            (7, 2): interviews,
            (8, 2): offers,
        }
        for (r, c), val in stat_map.items():
            try:
                ds.cell(row=r, column=c).value = val
            except Exception:
                pass

    wb.save(EXCEL_FILE)
    print(f"  Updated Excel → {EXCEL_FILE} ({len(contacts)} rows written)")


def main():
    print("\n🔍 Job Search Seeder — Montrez Cox")
    print("=" * 45)

    criteria = load_criteria()
    print(f"  Roles:    {criteria.get('roles')}")
    print(f"  Criteria: {criteria.get('worktype')} | min {criteria.get('minsalary')}")

    groq_key   = get_env_key("GROQ_API_KEY")
    gemini_key = get_env_key("GEMINI_API_KEY")

    leads = None

    # 1. Try Groq (free, fast, generous quota)
    if groq_key:
        try:
            leads = generate_job_leads_groq(criteria, groq_key, count=25)
        except Exception as e:
            print(f"  ⚠️  Groq error: {e} — trying Gemini...")

    # 2. Try Gemini
    if leads is None and gemini_key:
        try:
            leads = generate_job_leads(criteria, gemini_key, count=25)
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", errors="replace")
            if "429" in str(e.code) or "quota" in detail.lower():
                print("  ⚠️  Gemini quota exceeded.")
            else:
                print(f"  ⚠️  Gemini error ({e.code}).")
        except Exception as e:
            print(f"  ⚠️  Gemini unavailable: {e}")

    # 3. Built-in fallback
    if leads is None:
        print("  Using built-in job leads (add GROQ_API_KEY to .env for AI-generated leads).")
        leads = FALLBACK_LEADS

    # Build contact records
    contacts = build_contacts(leads)

    # Save contacts.json
    print("\n📁 Saving contacts.json...")
    save_contacts(contacts)

    # Update Excel
    print("\n📊 Updating Excel...")
    update_excel(contacts)

    print(f"\n✅ Done! {len(contacts)} job leads populated.")
    print(f"   contacts.json → {DATA_FILE}")
    print(f"   Excel         → {EXCEL_FILE}")
    print(f"\n   Open the dashboard to see the data:")
    print(f"   python3 server.py   (then go to http://localhost:8765)")


if __name__ == "__main__":
    main()
